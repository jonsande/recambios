from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any
from zipfile import BadZipFile

from django.conf import settings
from django.contrib.auth.models import AbstractBaseUser
from django.db import IntegrityError, transaction
from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from apps.catalog.models import (
    Brand,
    Category,
    Condition,
    PartNumber,
    PartNumberType,
    Product,
    normalize_part_number,
)

from .models import SupplierImport, SupplierImportRow
from .schema import (
    CANONICAL_IMPORT_COLUMNS,
    PART_NUMBER_COLUMNS,
    PART_NUMBER_REQUIRED_COLUMNS,
    validate_template_headers,
)

CLEAR_VALUE = "__CLEAR__"


class ImportProcessingError(Exception):
    pass


@dataclass
class ImportProcessingSummary:
    total_rows: int = 0
    successful_rows: int = 0
    failed_rows: int = 0
    skipped_rows: int = 0


def run_supplier_import(
    import_record: SupplierImport,
    requested_by: AbstractBaseUser,
) -> SupplierImport:
    import_record = SupplierImport.objects.select_related("supplier").get(pk=import_record.pk)
    import_record.import_status = SupplierImport.ImportStatus.PROCESSING
    import_record.started_at = timezone.now()
    import_record.finished_at = None
    import_record.total_rows = 0
    import_record.successful_rows = 0
    import_record.failed_rows = 0
    import_record.processing_notes = ""
    import_record.save(
        update_fields=[
            "import_status",
            "started_at",
            "finished_at",
            "total_rows",
            "successful_rows",
            "failed_rows",
            "processing_notes",
            "updated_at",
        ]
    )
    import_record.rows.all().delete()

    workbook = None
    try:
        workbook, worksheet, normalized_headers = _load_worksheet(import_record)
        header_result = validate_template_headers(normalized_headers)
        if header_result.duplicate_columns:
            duplicates = ", ".join(header_result.duplicate_columns)
            raise ImportProcessingError(f"Duplicate columns are not allowed: {duplicates}.")
        if header_result.missing_required:
            missing = ", ".join(header_result.missing_required)
            raise ImportProcessingError(f"Missing required columns: {missing}.")
        if "part_numbers" in workbook.sheetnames:
            reference_header_row = next(
                workbook["part_numbers"].iter_rows(min_row=1, max_row=1, values_only=True),
                None,
            )
            if reference_header_row:
                reference_headers = [
                    str(value).strip().lower() if value is not None else ""
                    for value in reference_header_row
                ]
                _validate_part_number_headers(reference_headers)
    except ImportProcessingError as exc:
        if workbook is not None:
            workbook.close()
        return _mark_import_failed(import_record, str(exc))

    summary = ImportProcessingSummary()
    import_warnings: list[str] = []
    successful_product_ids: set[int] = set()
    if header_result.unknown_columns:
        unknown = ", ".join(header_result.unknown_columns)
        import_warnings.append(f"Unknown columns ignored: {unknown}.")

    try:
        row_headers = normalized_headers
        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            summary.total_rows += 1
            raw_payload = _build_raw_payload(row_headers, row_values)
            if _is_empty_payload(raw_payload):
                SupplierImportRow.objects.create(
                    supplier_import=import_record,
                    row_number=row_number,
                    raw_payload=raw_payload,
                    processing_status=SupplierImportRow.ProcessingStatus.SKIPPED,
                    error_message="Empty row skipped.",
                )
                summary.skipped_rows += 1
                continue

            try:
                with transaction.atomic():
                    product, row_messages = _process_data_row(
                        import_record=import_record,
                        raw_payload=raw_payload,
                    )
                    SupplierImportRow.objects.create(
                        supplier_import=import_record,
                        row_number=row_number,
                        raw_payload=raw_payload,
                        processing_status=SupplierImportRow.ProcessingStatus.SUCCESS,
                        linked_product=product,
                        error_message="; ".join(row_messages),
                    )
                summary.successful_rows += 1
                successful_product_ids.add(product.pk)
            except ImportProcessingError as exc:
                SupplierImportRow.objects.create(
                    supplier_import=import_record,
                    row_number=row_number,
                    raw_payload=raw_payload,
                    processing_status=SupplierImportRow.ProcessingStatus.ERROR,
                    error_message=str(exc),
                )
                summary.failed_rows += 1
            except IntegrityError as exc:
                SupplierImportRow.objects.create(
                    supplier_import=import_record,
                    row_number=row_number,
                    raw_payload=raw_payload,
                    processing_status=SupplierImportRow.ProcessingStatus.ERROR,
                    error_message=f"Database integrity error: {exc}",
                )
                summary.failed_rows += 1
        reference_sheet = (
            workbook["part_numbers"] if "part_numbers" in workbook.sheetnames else None
        )
        if reference_sheet is not None:
            _process_part_number_sheet(
                import_record=import_record,
                worksheet=reference_sheet,
                row_number_offset=worksheet.max_row,
                successful_product_ids=successful_product_ids,
                summary=summary,
            )
    except Exception as exc:
        return _mark_import_failed(import_record, f"Unexpected processing error: {exc}")
    finally:
        if workbook is not None:
            workbook.close()

    return _finalize_import(import_record, summary, import_warnings)


def _load_worksheet(import_record: SupplierImport):
    if not import_record.original_file:
        raise ImportProcessingError("Missing file: upload a .xlsx file before processing.")
    file_name = import_record.original_file.name.lower()
    if not file_name.endswith(".xlsx"):
        raise ImportProcessingError("Only .xlsx files are supported in v1.")
    max_file_size = getattr(settings, "SUPPLIER_IMPORT_MAX_FILE_SIZE", 10 * 1024 * 1024)
    if import_record.original_file.size > max_file_size:
        raise ImportProcessingError(
            _("El archivo supera el tamaño máximo permitido de %(size)s bytes.")
            % {"size": max_file_size}
        )

    try:
        import_record.original_file.open("rb")
        workbook = load_workbook(import_record.original_file, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ImportProcessingError(f"Invalid XLSX file: {exc}") from exc

    if "products_import" not in workbook.sheetnames:
        workbook.close()
        raise ImportProcessingError(_("Falta la hoja obligatoria 'products_import'."))
    worksheet = workbook["products_import"]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ImportProcessingError("The worksheet is empty. Missing header row.")
    normalized_headers = [
        str(value).strip().lower() if value is not None else "" for value in header_row
    ]
    if not any(normalized_headers):
        raise ImportProcessingError("The header row is empty.")
    max_rows = getattr(settings, "SUPPLIER_IMPORT_MAX_ROWS", 10_000)
    product_rows = max(worksheet.max_row - 1, 0)
    reference_rows = (
        max(workbook["part_numbers"].max_row - 1, 0) if "part_numbers" in workbook.sheetnames else 0
    )
    if product_rows + reference_rows > max_rows:
        workbook.close()
        raise ImportProcessingError(
            _("El archivo supera el máximo de %(rows)s filas de datos.") % {"rows": max_rows}
        )
    return workbook, worksheet, normalized_headers


def _mark_import_failed(import_record: SupplierImport, notes: str) -> SupplierImport:
    import_record.import_status = SupplierImport.ImportStatus.FAILED
    import_record.finished_at = timezone.now()
    import_record.processing_notes = notes
    import_record.save(
        update_fields=["import_status", "finished_at", "processing_notes", "updated_at"]
    )
    return import_record


def _finalize_import(
    import_record: SupplierImport,
    summary: ImportProcessingSummary,
    warnings: list[str],
) -> SupplierImport:
    notes: list[str] = []
    notes.extend(warnings)
    if summary.skipped_rows:
        notes.append(f"Skipped empty rows: {summary.skipped_rows}.")

    if summary.total_rows == 0:
        import_record.import_status = SupplierImport.ImportStatus.FAILED
        notes.append("No data rows found to process.")
    elif summary.successful_rows + summary.failed_rows == 0:
        import_record.import_status = SupplierImport.ImportStatus.FAILED
        notes.append(_("No se encontró ninguna fila útil para importar."))
    elif summary.failed_rows == 0 and not warnings and summary.skipped_rows == 0:
        import_record.import_status = SupplierImport.ImportStatus.COMPLETED
    else:
        import_record.import_status = SupplierImport.ImportStatus.COMPLETED_WITH_ERRORS

    import_record.total_rows = summary.total_rows
    import_record.successful_rows = summary.successful_rows
    import_record.failed_rows = summary.failed_rows
    import_record.finished_at = timezone.now()
    import_record.processing_notes = "\n".join(notes)
    import_record.save(
        update_fields=[
            "import_status",
            "total_rows",
            "successful_rows",
            "failed_rows",
            "finished_at",
            "processing_notes",
            "updated_at",
        ]
    )
    return import_record


def _process_data_row(
    import_record: SupplierImport,
    raw_payload: dict[str, Any],
) -> tuple[Product, list[str]]:
    title = _clean_text(raw_payload.get("product_title"))
    brand_name = _clean_text(raw_payload.get("brand_name"))
    category_slug = _clean_text(raw_payload.get("category_slug"))
    condition_code = _clean_text(raw_payload.get("condition_code"))
    sku = _clean_text(raw_payload.get("oe_code"))
    supplier_product_code = _clean_text(raw_payload.get("supplier_product_code"))

    row_messages: list[str] = []
    if not title:
        raise ImportProcessingError("Missing required value: product_title.")
    if not category_slug:
        raise ImportProcessingError("Missing required value: category_slug.")
    if not condition_code:
        raise ImportProcessingError("Missing required value: condition_code.")
    if not sku and not supplier_product_code:
        raise ImportProcessingError("Each row must include OE_code or supplier_product_code.")

    target_product = _match_existing_product(
        import_record=import_record,
        sku=sku,
        supplier_product_code=supplier_product_code,
    )

    if target_product is None and not sku:
        raise ImportProcessingError(
            "OE_code is required to create a new product when no existing match is found."
        )

    brand = _resolve_brand(brand_name) if brand_name and brand_name != CLEAR_VALUE else None
    category = _resolve_category(category_slug)
    condition = Condition.objects.filter(code__iexact=condition_code, is_active=True).first()
    if not condition:
        raise ImportProcessingError(
            f"Unknown condition_code '{condition_code}'. Create it first in canonical conditions."
        )

    last_known_price = _parse_decimal(
        raw_payload.get("last_known_price"),
        field_name="last_known_price",
    )
    dimensions = {
        name: _parse_decimal(raw_payload.get(name), field_name=name)
        for name in ("weight", "length", "width", "height")
    }
    if any(value is not None and value < 0 for value in dimensions.values()):
        raise ImportProcessingError("weight and dimensions cannot be negative.")
    currency = _clean_text(raw_payload.get("currency")).upper()
    unit_of_sale = _clean_text(raw_payload.get("unit_of_sale"))
    quantity = _parse_positive_integer(raw_payload.get("quantity"), field_name="quantity")
    unit_of_quantity = _clean_text(raw_payload.get("unit_of_quantity"))
    short_description = _clean_text(raw_payload.get("short_description"))
    long_description = _clean_text(raw_payload.get("long_description"))
    if target_product is None:
        supplier_product_code = (
            "" if supplier_product_code == CLEAR_VALUE else supplier_product_code
        )
        short_description = "" if short_description == CLEAR_VALUE else short_description
        long_description = "" if long_description == CLEAR_VALUE else long_description

    if currency and len(currency) != 3:
        raise ImportProcessingError("currency must be a 3-letter code.")

    if target_product is None:
        product = Product(
            supplier=import_record.supplier,
            sku=sku,
            supplier_product_code=supplier_product_code or None,
            title=title,
            short_description=short_description,
            long_description=long_description,
            brand=brand,
            category=category,
            condition=condition,
            price_visibility_mode=Product.PriceVisibilityMode.VISIBLE_INFO,
            last_known_price=last_known_price,
            currency=currency or "EUR",
            unit_of_sale=unit_of_sale or "unit",
            quantity=quantity or 1,
            unit_of_quantity=unit_of_quantity or "Pcs",
            is_active=False,
            featured=False,
            **dimensions,
        )
        product.save()
    else:
        product = target_product
        product.title = title
        if brand_name == CLEAR_VALUE:
            product.brand = None
        elif brand_name:
            product.brand = brand
        product.category = category
        product.condition = condition
        product.supplier = import_record.supplier

        if sku and sku != product.sku:
            conflict = Product.objects.exclude(pk=product.pk).filter(sku=sku).exists()
            if conflict:
                raise ImportProcessingError(f"sku '{sku}' already exists in another product.")
            product.sku = sku

        if supplier_product_code:
            product.supplier_product_code = (
                None if supplier_product_code == CLEAR_VALUE else supplier_product_code
            )
        if short_description == CLEAR_VALUE:
            product.short_description = ""
        elif short_description:
            product.short_description = short_description
        if long_description == CLEAR_VALUE:
            product.long_description = ""
        elif long_description:
            product.long_description = long_description
        if _is_clear(raw_payload.get("last_known_price")):
            product.last_known_price = None
        elif last_known_price is not None:
            product.last_known_price = last_known_price
        if currency:
            product.currency = currency
        if unit_of_sale:
            product.unit_of_sale = unit_of_sale
        if quantity is not None:
            product.quantity = quantity
        if unit_of_quantity:
            product.unit_of_quantity = unit_of_quantity
        for field_name, value in dimensions.items():
            if _is_clear(raw_payload.get(field_name)):
                setattr(product, field_name, None)
            elif value is not None:
                setattr(product, field_name, value)

        product.publication_status = Product.PublicationStatus.DRAFT
        product.published_at = None

        product.save()

    return product, row_messages


def _match_existing_product(
    import_record: SupplierImport,
    sku: str,
    supplier_product_code: str,
) -> Product | None:
    sku_match = Product.objects.filter(sku=sku).first() if sku else None
    supplier_match = (
        Product.objects.filter(
            supplier=import_record.supplier,
            supplier_product_code=supplier_product_code,
        ).first()
        if supplier_product_code
        else None
    )

    if sku_match and sku_match.supplier_id != import_record.supplier_id:
        raise ImportProcessingError(
            f"sku '{sku}' belongs to supplier '{sku_match.supplier.code}', "
            f"not '{import_record.supplier.code}'."
        )
    if sku_match and supplier_match and sku_match.pk != supplier_match.pk:
        raise ImportProcessingError("sku and supplier_product_code point to different products.")

    matched_product = sku_match or supplier_match
    if matched_product and matched_product.supplier_id != import_record.supplier_id:
        raise ImportProcessingError(
            "Matched product belongs to another supplier and cannot be updated in this import."
        )
    return matched_product


def _resolve_brand(name: str) -> Brand:
    brand = Brand.objects.filter(name__iexact=name, is_active=True).first()
    if brand is None:
        raise ImportProcessingError(f"Unknown or inactive brand_name '{name}'.")
    return brand


def _resolve_category(category_slug: str) -> Category:
    category = Category.objects.filter(slug=category_slug, is_active=True).first()
    if category is None:
        raise ImportProcessingError(f"Unknown or inactive category_slug '{category_slug}'.")
    return category


def _validate_part_number_headers(headers: list[str]) -> None:
    duplicate_columns = {header for header in headers if header and headers.count(header) > 1}
    if duplicate_columns:
        raise ImportProcessingError(
            f"Duplicate columns in part_numbers: {', '.join(sorted(duplicate_columns))}."
        )
    missing = [column for column in PART_NUMBER_REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ImportProcessingError(
            f"Missing required columns in part_numbers: {', '.join(missing)}."
        )


def _process_part_number_sheet(
    *,
    import_record: SupplierImport,
    worksheet,
    row_number_offset: int,
    successful_product_ids: set[int],
    summary: ImportProcessingSummary,
) -> None:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return
    headers = [str(value).strip().lower() if value is not None else "" for value in header_row]
    _validate_part_number_headers(headers)

    for sheet_row, row_values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        trace_row = row_number_offset + sheet_row
        payload = _build_raw_payload_for_columns(headers, row_values, PART_NUMBER_COLUMNS)
        payload["_sheet"] = "part_numbers"
        payload["_sheet_row"] = sheet_row
        summary.total_rows += 1
        if _is_empty_payload(payload, ignored_keys={"_sheet", "_sheet_row"}):
            SupplierImportRow.objects.create(
                supplier_import=import_record,
                row_number=trace_row,
                raw_payload=payload,
                processing_status=SupplierImportRow.ProcessingStatus.SKIPPED,
                error_message=_("Fila vacía omitida en part_numbers."),
            )
            summary.skipped_rows += 1
            continue
        try:
            with transaction.atomic():
                product = _process_part_number_row(
                    import_record=import_record,
                    payload=payload,
                    successful_product_ids=successful_product_ids,
                )
                SupplierImportRow.objects.create(
                    supplier_import=import_record,
                    row_number=trace_row,
                    raw_payload=payload,
                    processing_status=SupplierImportRow.ProcessingStatus.SUCCESS,
                    linked_product=product,
                )
            summary.successful_rows += 1
        except (ImportProcessingError, IntegrityError) as exc:
            SupplierImportRow.objects.create(
                supplier_import=import_record,
                row_number=trace_row,
                raw_payload=payload,
                processing_status=SupplierImportRow.ProcessingStatus.ERROR,
                error_message=str(exc),
            )
            summary.failed_rows += 1


def _process_part_number_row(
    *, import_record: SupplierImport, payload: dict[str, Any], successful_product_ids: set[int]
) -> Product:
    sku = _clean_text(payload.get("oe_code"))
    number = _clean_text(payload.get("number"))
    type_code = _clean_text(payload.get("type_code"))
    if not sku or not number or not type_code:
        raise ImportProcessingError(
            "part_numbers requires OE_code, number and type_code in every row."
        )
    product = Product.objects.filter(supplier=import_record.supplier, sku=sku).first()
    if product is None or product.pk not in successful_product_ids:
        raise ImportProcessingError(f"OE_code '{sku}' was not imported successfully in this file.")
    part_number_type = PartNumberType.objects.filter(code__iexact=type_code, is_active=True).first()
    if part_number_type is None:
        raise ImportProcessingError(f"Unknown or inactive type_code '{type_code}'.")
    brand_name = _clean_text(payload.get("brand_name"))
    brand = _resolve_brand(brand_name) if brand_name else None
    normalized = normalize_part_number(number)
    if not normalized:
        raise ImportProcessingError("number must contain letters or digits.")
    part_number = PartNumber.objects.filter(
        product=product,
        number_normalized=normalized,
        part_number_type=part_number_type,
    ).first()
    if part_number is None:
        PartNumber.objects.create(
            product=product,
            number_raw=number,
            part_number_type=part_number_type,
            brand=brand,
            notes=_clean_text(payload.get("notes")),
            is_primary=False,
        )
    else:
        if part_number.is_primary:
            raise ImportProcessingError(
                "The additional reference matches the primary SKU reference."
            )
        part_number.number_raw = number
        part_number.brand = brand
        part_number.notes = _clean_text(payload.get("notes")) or part_number.notes
        part_number.is_primary = False
        part_number.save()
    return product


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_decimal(value: Any, *, field_name: str) -> Decimal | None:
    text_value = _clean_text(value)
    if not text_value or text_value == CLEAR_VALUE:
        return None
    try:
        return Decimal(text_value)
    except (InvalidOperation, ValueError) as exc:
        raise ImportProcessingError(
            f"Invalid decimal value for {field_name}: '{text_value}'."
        ) from exc


def _parse_positive_integer(value: Any, *, field_name: str) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raise ImportProcessingError(f"Invalid integer value for {field_name}: '{value}'.")

    text_value = _clean_text(value)
    if not text_value:
        return None

    try:
        parsed_value = Decimal(text_value)
    except (InvalidOperation, ValueError) as exc:
        raise ImportProcessingError(
            f"Invalid integer value for {field_name}: '{text_value}'."
        ) from exc

    if parsed_value != parsed_value.to_integral_value():
        raise ImportProcessingError(f"Invalid integer value for {field_name}: '{text_value}'.")

    int_value = int(parsed_value)
    if int_value < 1:
        raise ImportProcessingError(f"{field_name} must be greater than or equal to 1.")
    return int_value


def _serialize_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _build_raw_payload(headers: list[str], row_values: tuple[Any, ...]) -> dict[str, Any]:
    return _build_raw_payload_for_columns(headers, row_values, CANONICAL_IMPORT_COLUMNS)


def _build_raw_payload_for_columns(
    headers: list[str], row_values: tuple[Any, ...], known_columns: tuple[str, ...]
) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        cell_value = row_values[index] if index < len(row_values) else None
        payload[header] = _serialize_cell_value(cell_value)
    for known_column in known_columns:
        payload.setdefault(known_column, None)
    return payload


def _is_empty_payload(raw_payload: dict[str, Any], *, ignored_keys: set[str] | None = None) -> bool:
    ignored_keys = ignored_keys or set()
    for key, value in raw_payload.items():
        if key in ignored_keys:
            continue
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return False
    return True


def _is_clear(value: Any) -> bool:
    return _clean_text(value).upper() == CLEAR_VALUE

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from apps.catalog.models import Brand, Category, Condition, PartNumberType

from .schema import CANONICAL_IMPORT_COLUMNS, PART_NUMBER_COLUMNS


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")


def _configure_data_sheet(sheet, columns: tuple[str, ...]) -> None:
    sheet.append(["OE_code" if column == "oe_code" else column for column in columns])
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    for index, column_name in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(16, len(column_name) + 2)


def _add_list_validation(sheet, column_name: str, formula: str, columns: tuple[str, ...]) -> None:
    column_index = columns.index(column_name) + 1
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Select a value from the list."
    validation.errorTitle = "Invalid value"
    sheet.add_data_validation(validation)
    letter = get_column_letter(column_index)
    validation.add(f"{letter}2:{letter}10001")


def build_supplier_import_template_xlsx() -> bytes:
    workbook = Workbook()
    products_sheet = workbook.active
    products_sheet.title = "products_import"
    _configure_data_sheet(products_sheet, CANONICAL_IMPORT_COLUMNS)

    references_sheet = workbook.create_sheet("part_numbers")
    _configure_data_sheet(references_sheet, PART_NUMBER_COLUMNS)

    instructions = workbook.create_sheet("instructions")
    instructions.append(["Supplier import template v2"])
    instructions.append(["Sheet", "Field", "Required", "Description / format"])
    rows = [
        ("products_import", "product_title", "yes", "Product title."),
        (
            "products_import",
            "OE_code",
            "new: yes",
            "Primary manufacturer OE/OES reference; this is not an arbitrary internal SKU.",
        ),
        (
            "products_import",
            "supplier_product_code",
            "alternative",
            "Supplier's own product code; it can be used to locate products for updates.",
        ),
        ("products_import", "category_slug", "yes", "Slug of an existing active category."),
        (
            "products_import",
            "brand_name",
            "no",
            "Existing active brand. Blank preserves the value; __CLEAR__ removes it.",
        ),
        (
            "products_import",
            "condition_code",
            "yes",
            "Code of an existing active product condition.",
        ),
        (
            "products_import",
            "last_known_price",
            "no",
            "Unit price excluding VAT; __CLEAR__ removes the price.",
        ),
        (
            "products_import",
            "weight/length/width/height",
            "no",
            "Non-negative decimals; __CLEAR__ removes the value.",
        ),
        (
            "part_numbers",
            "OE_code",
            "yes",
            "Primary OE code of a successfully imported product in this file.",
        ),
        (
            "part_numbers",
            "number",
            "yes",
            "Additional part number; it never replaces the primary reference.",
        ),
        (
            "part_numbers",
            "type_code",
            "yes",
            "Active part-number type, for example OEM, OES or AIM.",
        ),
        ("part_numbers", "brand_name", "no", "Existing active brand."),
    ]
    for row in rows:
        instructions.append(row)
    instructions.append([])
    instructions.append(["Rule", "Blank cells preserve existing data when updating a product."])
    instructions.append(
        [
            "Rule",
            "Every created or modified product is saved as a draft and must be reviewed before "
            "publication. New products are created inactive, not featured, with 21% VAT and a "
            "visible price.",
        ]
    )
    instructions.append(
        [
            "Rule",
            "Brands, categories, conditions and part-number types are not created automatically.",
        ]
    )
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 28
    instructions.column_dimensions["C"].width = 16
    instructions.column_dimensions["D"].width = 90

    example = workbook.create_sheet("example")
    example.append(
        ["OE_code" if column == "oe_code" else column for column in CANONICAL_IMPORT_COLUMNS]
    )
    example.append(
        [
            "12V starter motor",
            "Bosch",
            "motores-arranque",
            "new",
            "0986012345",
            "PROV-001",
            "125.50",
            "EUR",
            "unit",
            1,
            "Pcs",
            "4.250",
            "30",
            "18",
            "20",
            "New starter motor",
            "Full product description",
        ]
    )
    _style_header(example)

    values = workbook.create_sheet("allowed_values")
    value_columns = {
        "category_slug": list(
            Category.objects.filter(is_active=True).order_by("slug").values_list("slug", flat=True)
        ),
        "brand_name": list(
            Brand.objects.filter(is_active=True).order_by("name").values_list("name", flat=True)
        ),
        "condition_code": list(
            Condition.objects.filter(is_active=True).order_by("code").values_list("code", flat=True)
        ),
        "type_code": list(
            PartNumberType.objects.filter(is_active=True)
            .order_by("code")
            .values_list("code", flat=True)
        ),
    }
    range_names = {
        "category_slug": "allowed_categories",
        "brand_name": "allowed_brands",
        "condition_code": "allowed_conditions",
        "type_code": "allowed_part_number_types",
    }
    for column_index, (heading, allowed_values) in enumerate(value_columns.items(), start=1):
        values.cell(row=1, column=column_index, value=heading)
        for row_index, value in enumerate(allowed_values, start=2):
            values.cell(row=row_index, column=column_index, value=value)
        values.column_dimensions[get_column_letter(column_index)].width = max(20, len(heading) + 2)
    _style_header(values)

    for column_index, (heading, allowed_values) in enumerate(value_columns.items(), start=1):
        if not allowed_values:
            continue
        end_row = len(allowed_values) + 1
        letter = get_column_letter(column_index)
        range_name = range_names[heading]
        workbook.defined_names.add(
            DefinedName(
                range_name,
                attr_text=f"'allowed_values'!${letter}$2:${letter}${end_row}",
            )
        )
        formula = f"={range_name}"
        if heading == "type_code":
            _add_list_validation(references_sheet, heading, formula, PART_NUMBER_COLUMNS)
            continue
        _add_list_validation(products_sheet, heading, formula, CANONICAL_IMPORT_COLUMNS)
        if heading == "brand_name":
            _add_list_validation(references_sheet, heading, formula, PART_NUMBER_COLUMNS)

    content = BytesIO()
    workbook.save(content)
    workbook.close()
    return content.getvalue()

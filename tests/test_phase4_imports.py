from __future__ import annotations

from io import BytesIO

import pytest
from django.contrib.admin.sites import AdminSite
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, override_settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.catalog.admin import ProductAdmin
from apps.catalog.models import Brand, Category, Condition, PartNumber, PartNumberType, Product
from apps.imports.admin import SupplierImportAdmin
from apps.imports.models import SupplierImport, SupplierImportRow
from apps.imports.schema import CANONICAL_IMPORT_COLUMNS, validate_template_headers
from apps.imports.services import run_supplier_import
from apps.imports.template_builder import build_supplier_import_template_xlsx
from apps.suppliers.models import Supplier
from apps.users.roles import ROLE_INTERNAL_STAFF, ROLE_RESTRICTED_SUPPLIER


def make_supplier(code: str = "SUP-IMP") -> Supplier:
    return Supplier.objects.create(
        name=f"Supplier {code}",
        slug=f"supplier-{code.lower()}",
        code=code,
    )


def make_condition(code: str = "new", name: str = "Nuevo", slug: str = "new") -> Condition:
    return Condition.objects.create(code=code, name=name, slug=slug)


def make_staff_user(django_user_model, username: str):
    return django_user_model.objects.create_user(
        username=username,
        email=f"{username}@example.com",
        password="pass1234",
        is_staff=True,
    )


def build_import_file(
    headers: list[str],
    rows: list[list[object]],
    reference_rows: list[list[object]] | None = None,
) -> SimpleUploadedFile:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "products_import"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    if reference_rows is not None:
        reference_sheet = workbook.create_sheet("part_numbers")
        reference_sheet.append(["OE_code", "number", "type_code", "brand_name", "notes"])
        for row in reference_rows:
            reference_sheet.append(row)

    content = BytesIO()
    workbook.save(content)
    workbook.close()
    return SimpleUploadedFile(
        "supplier_import.xlsx",
        content.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def canonical_row(**values: object) -> list[object]:
    row = {column: "" for column in CANONICAL_IMPORT_COLUMNS}
    row.update(values)
    return [row[column] for column in CANONICAL_IMPORT_COLUMNS]


@pytest.mark.django_db
def test_valid_template_headers_are_recognized() -> None:
    header_result = validate_template_headers(list(CANONICAL_IMPORT_COLUMNS))

    assert header_result.is_valid
    assert not header_result.missing_required
    assert not header_result.duplicate_columns


@pytest.mark.django_db
def test_template_builder_generates_canonical_headers() -> None:
    Brand.objects.create(name="Dropdown Brand", slug="dropdown-brand")
    Category.objects.create(name="Dropdown Category", slug="dropdown-category")
    make_condition("dropdown-condition", "Dropdown condition", "dropdown-condition")
    file_content = build_supplier_import_template_xlsx()
    workbook = load_workbook(BytesIO(file_content), data_only=True)
    sheet = workbook["products_import"]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    normalized_headers = tuple(
        str(value).strip().lower() for value in header_row if value is not None
    )
    assert normalized_headers == CANONICAL_IMPORT_COLUMNS
    assert "OE_code" in header_row
    assert "category_name" not in header_row
    assert "price_visibility_mode" not in header_row
    assert "product_vat_rate" not in header_row
    assert "is_active" not in header_row
    assert "featured" not in header_row
    assert header_row[-2:] == ("short_description", "long_description")
    assert {"part_numbers", "instructions", "example", "allowed_values"}.issubset(
        workbook.sheetnames
    )
    assert not {"instrucciones", "ejemplo", "valores_permitidos"}.intersection(workbook.sheetnames)
    instructions = workbook["instructions"]
    assert instructions["A1"].value == "Supplier import template v2"
    assert instructions["A2"].value == "Sheet"
    assert instructions["D2"].value == "Description / format"
    assert {
        "allowed_brands",
        "allowed_categories",
        "allowed_conditions",
        "allowed_part_number_types",
    }.issubset(workbook.defined_names)
    product_formulas = {
        validation.formula1
        for validation in workbook["products_import"].data_validations.dataValidation
    }
    reference_formulas = {
        validation.formula1
        for validation in workbook["part_numbers"].data_validations.dataValidation
    }
    assert {"=allowed_brands", "=allowed_categories", "=allowed_conditions"}.issubset(
        product_formulas
    )
    assert {"=allowed_brands", "=allowed_part_number_types"}.issubset(reference_formulas)
    workbook.close()


@pytest.mark.django_db
def test_v2_imports_commercial_fields_and_additional_part_numbers(django_user_model) -> None:
    supplier = make_supplier("SUP-V2")
    user = make_staff_user(django_user_model, "import_user_v2")
    category = Category.objects.create(name="Motores", slug="motores")
    brand = Brand.objects.create(name="Bosch", slug="bosch")
    make_condition("new", "Nuevo", "new")
    PartNumberType.objects.get_or_create(code="OES", defaults={"name": "OES"})
    row = canonical_row(
        product_title="Producto v2",
        brand_name=brand.name,
        condition_code="new",
        oe_code="OE-V2-1",
        supplier_product_code="SUP-V2-1",
        category_slug=category.slug,
        last_known_price="99.95",
        weight="2.500",
        length="20.000",
        width="10.000",
        height="5.000",
    )
    import_file = build_import_file(
        list(CANONICAL_IMPORT_COLUMNS),
        [row],
        reference_rows=[["OE-V2-1", "ALT-123", "OES", "Bosch", "Equivalente"]],
    )
    record = SupplierImport.objects.create(
        supplier=supplier, uploaded_by=user, original_file=import_file
    )

    result = run_supplier_import(record, user)
    product = Product.objects.get(sku="OE-V2-1")

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED
    assert product.publication_status == Product.PublicationStatus.DRAFT
    assert product.product_vat_rate == 21
    assert product.price_visibility_mode == Product.PriceVisibilityMode.VISIBLE_INFO
    assert not product.is_active
    assert not product.featured
    assert product.weight == 2.5
    assert PartNumber.objects.filter(
        product=product, number_normalized="ALT123", part_number_type__code="OES"
    ).exists()


@pytest.mark.django_db
def test_import_updates_published_product_as_draft_and_supports_clear(django_user_model) -> None:
    supplier = make_supplier("SUP-CLEAR")
    user = make_staff_user(django_user_model, "import_user_clear")
    category = Category.objects.create(name="Motores", slug="motores-clear")
    condition = make_condition("new", "Nuevo", "new")
    brand = Brand.objects.create(name="Bosch", slug="bosch-clear")
    product = Product.objects.create(
        supplier=supplier,
        sku="OE-CLEAR-1",
        title="Anterior",
        category=category,
        condition=condition,
        brand=brand,
        last_known_price="50",
        weight="3",
        product_vat_rate="10",
        price_visibility_mode=Product.PriceVisibilityMode.HIDDEN,
        is_active=True,
        featured=True,
        publication_status=Product.PublicationStatus.PUBLISHED,
        published_at=timezone.now(),
    )
    row = canonical_row(
        product_title="Actualizado",
        brand_name="__CLEAR__",
        condition_code="new",
        oe_code=product.sku,
        category_slug=category.slug,
        last_known_price="__CLEAR__",
        weight="__CLEAR__",
    )
    record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=build_import_file(
            list(CANONICAL_IMPORT_COLUMNS),
            [row],
        ),
    )

    result = run_supplier_import(record, user)
    product.refresh_from_db()

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED
    assert product.publication_status == Product.PublicationStatus.DRAFT
    assert product.published_at is None
    assert product.brand is None
    assert product.last_known_price is None
    assert product.weight is None
    assert product.product_vat_rate == 10
    assert product.price_visibility_mode == Product.PriceVisibilityMode.HIDDEN
    assert product.is_active
    assert product.featured


@pytest.mark.django_db
def test_import_rejects_unknown_master_data_without_creating_it(django_user_model) -> None:
    supplier = make_supplier("SUP-MASTER")
    user = make_staff_user(django_user_model, "import_user_master")
    make_condition("new", "Nuevo", "new")
    record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=build_import_file(
            ["product_title", "brand_name", "category_slug", "condition_code", "OE_code"],
            [["Producto", "Marca inexistente", "categoria-inexistente", "new", "OE-X"]],
        ),
    )

    result = run_supplier_import(record, user)

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED_WITH_ERRORS
    assert not Brand.objects.filter(name="Marca inexistente").exists()
    assert not Category.objects.filter(name="Categoría inexistente").exists()


@pytest.mark.django_db
def test_import_with_only_empty_rows_fails(django_user_model) -> None:
    supplier = make_supplier("SUP-EMPTY")
    user = make_staff_user(django_user_model, "import_user_empty")
    record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=build_import_file(list(CANONICAL_IMPORT_COLUMNS), [[""]]),
    )

    result = run_supplier_import(record, user)

    assert result.import_status == SupplierImport.ImportStatus.FAILED
    assert result.successful_rows == 0


@pytest.mark.django_db
@override_settings(SUPPLIER_IMPORT_MAX_ROWS=1)
def test_import_enforces_configured_row_limit(django_user_model) -> None:
    supplier = make_supplier("SUP-LIMIT")
    user = make_staff_user(django_user_model, "import_user_limit")
    record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=build_import_file(
            ["product_title", "category_slug", "condition_code", "OE_code"],
            [["Uno", "Cat", "new", "ONE"], ["Dos", "Cat", "new", "TWO"]],
        ),
    )

    result = run_supplier_import(record, user)

    assert result.import_status == SupplierImport.ImportStatus.FAILED
    assert "máximo de 1 filas" in result.processing_notes


@pytest.mark.django_db
def test_import_fails_when_required_columns_are_missing(django_user_model) -> None:
    supplier = make_supplier("SUP-MISS-COL")
    user = make_staff_user(django_user_model, "import_user_missing")
    import_file = build_import_file(
        headers=["product_title", "condition_code", "OE_code"],
        rows=[["Starter Motor", "new", "SKU-1"]],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)

    assert result.import_status == SupplierImport.ImportStatus.FAILED
    assert "Missing required columns" in result.processing_notes
    assert result.rows.count() == 0


@pytest.mark.django_db
def test_import_creates_product_with_blank_brand_name(django_user_model) -> None:
    supplier = make_supplier("SUP-BRAND-BLANK")
    user = make_staff_user(django_user_model, "import_user_brand_blank")
    make_condition("new", "Nuevo", "new")
    Category.objects.create(name="Alternator", slug="alternator")

    import_file = build_import_file(
        headers=list(CANONICAL_IMPORT_COLUMNS),
        rows=[
            canonical_row(
                product_title="Brand Optional Product",
                category_slug="alternator",
                condition_code="new",
                oe_code="SKU-BRAND-BLANK-1",
                supplier_product_code="SUP-BRAND-BLANK-1",
            )
        ],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)
    product = Product.objects.get(sku="SKU-BRAND-BLANK-1")

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED
    assert product.brand is None


@pytest.mark.django_db
def test_import_persists_quantity_fields_when_provided(django_user_model) -> None:
    supplier = make_supplier("SUP-QTY-IMP")
    user = make_staff_user(django_user_model, "import_user_quantity")
    make_condition("new", "Nuevo", "new")
    Category.objects.create(name="Alternator", slug="alternator")

    import_file = build_import_file(
        headers=list(CANONICAL_IMPORT_COLUMNS),
        rows=[
            canonical_row(
                product_title="Quantity Product",
                category_slug="alternator",
                condition_code="new",
                oe_code="SKU-QTY-IMPORT-1",
                supplier_product_code="SUP-QTY-IMPORT-1",
                unit_of_sale="pack",
                quantity="6",
                unit_of_quantity="pairs",
            )
        ],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)
    product = Product.objects.get(sku="SKU-QTY-IMPORT-1")

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED
    assert product.quantity == 6
    assert product.unit_of_quantity == "pairs"


@pytest.mark.django_db
def test_import_records_row_level_traceability(django_user_model) -> None:
    supplier = make_supplier("SUP-TRACE")
    user = make_staff_user(django_user_model, "import_user_trace")
    make_condition("new", "Nuevo", "new")
    Brand.objects.create(name="Bosch", slug="bosch")
    Brand.objects.create(name="Brembo", slug="brembo")
    Category.objects.create(name="Starter", slug="starter")
    Category.objects.create(name="Brakes", slug="brakes")
    import_file = build_import_file(
        headers=list(CANONICAL_IMPORT_COLUMNS),
        rows=[
            canonical_row(
                product_title="Starter Motor",
                brand_name="Bosch",
                category_slug="starter",
                condition_code="new",
                oe_code="SKU-TRACE-1",
                supplier_product_code="SP-1",
            ),
            canonical_row(
                product_title="Brake Pad",
                brand_name="Brembo",
                category_slug="brakes",
                condition_code="missing",
                oe_code="SKU-TRACE-2",
                supplier_product_code="SP-2",
            ),
        ],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED_WITH_ERRORS
    assert result.total_rows == 2
    assert result.successful_rows == 1
    assert result.failed_rows == 1

    success_row = SupplierImportRow.objects.get(supplier_import=result, row_number=2)
    error_row = SupplierImportRow.objects.get(supplier_import=result, row_number=3)
    assert success_row.processing_status == SupplierImportRow.ProcessingStatus.SUCCESS
    assert success_row.linked_product is not None
    assert success_row.raw_payload["oe_code"] == "SKU-TRACE-1"
    assert error_row.processing_status == SupplierImportRow.ProcessingStatus.ERROR
    assert "Unknown condition_code" in error_row.error_message


@pytest.mark.django_db
def test_import_updates_existing_product_by_sku(django_user_model) -> None:
    supplier = make_supplier("SUP-SKU-UPD")
    user = make_staff_user(django_user_model, "import_user_sku")
    brand = Brand.objects.create(name="Brand One", slug="brand-one")
    category = Category.objects.create(name="Category One", slug="category-one")
    condition = make_condition("new", "Nuevo", "new")
    product = Product.objects.create(
        supplier=supplier,
        supplier_product_code="SUP-SKU-UPD-1",
        sku="SKU-UPD-1",
        slug="sku-upd-1",
        title="Old title",
        brand=brand,
        category=category,
        condition=condition,
    )
    import_file = build_import_file(
        headers=list(CANONICAL_IMPORT_COLUMNS),
        rows=[
            canonical_row(
                product_title="Updated title",
                brand_name="Brand One",
                category_slug="category-one",
                condition_code="new",
                oe_code="SKU-UPD-1",
                supplier_product_code="SUP-SKU-UPD-1",
            )
        ],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)
    product.refresh_from_db()

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED
    assert product.title == "Updated title"
    assert Product.objects.filter(supplier=supplier).count() == 1


@pytest.mark.django_db
def test_import_preserves_existing_brand_when_brand_name_is_blank(django_user_model) -> None:
    supplier = make_supplier("SUP-BRAND-PRES")
    user = make_staff_user(django_user_model, "import_user_brand_preserve")
    brand = Brand.objects.create(name="Preserve Brand", slug="preserve-brand")
    category = Category.objects.create(name="Starter", slug="starter-pres")
    condition = make_condition("pres-new", "Nuevo Pres", "pres-new")
    product = Product.objects.create(
        supplier=supplier,
        supplier_product_code="SUP-BRAND-PRES-1",
        sku="SKU-BRAND-PRES-1",
        slug="sku-brand-pres-1",
        title="Old Brand Product",
        brand=brand,
        category=category,
        condition=condition,
    )

    import_file = build_import_file(
        headers=list(CANONICAL_IMPORT_COLUMNS),
        rows=[
            canonical_row(
                product_title="Updated Brand Product",
                category_slug="starter-pres",
                condition_code="pres-new",
                oe_code="SKU-BRAND-PRES-1",
                supplier_product_code="SUP-BRAND-PRES-1",
            )
        ],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)
    product.refresh_from_db()

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED
    assert product.title == "Updated Brand Product"
    assert product.brand_id == brand.id


@pytest.mark.django_db
def test_import_updates_by_supplier_product_code_when_sku_missing(django_user_model) -> None:
    supplier = make_supplier("SUP-SUPCODE-UPD")
    user = make_staff_user(django_user_model, "import_user_supcode")
    brand = Brand.objects.create(name="Valeo", slug="valeo")
    category = Category.objects.create(name="Alternator", slug="alternator")
    condition = make_condition("used", "Usado", "used")
    product = Product.objects.create(
        supplier=supplier,
        supplier_product_code="SUPCODE-42",
        sku="SKU-SUPCODE-42",
        slug="sku-supcode-42",
        title="Old title",
        brand=brand,
        category=category,
        condition=condition,
    )
    import_file = build_import_file(
        headers=list(CANONICAL_IMPORT_COLUMNS),
        rows=[
            canonical_row(
                product_title="Updated by supplier code",
                brand_name="Valeo",
                category_slug="alternator",
                condition_code="used",
                supplier_product_code="SUPCODE-42",
            )
        ],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)
    product.refresh_from_db()

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED
    assert product.title == "Updated by supplier code"


@pytest.mark.django_db
def test_import_errors_when_sku_and_supplier_code_conflict(django_user_model) -> None:
    supplier = make_supplier("SUP-CONFLICT")
    user = make_staff_user(django_user_model, "import_user_conflict")
    brand = Brand.objects.create(name="Brand C", slug="brand-c")
    category = Category.objects.create(name="Category C", slug="category-c")
    condition = make_condition("new", "Nuevo", "new")
    Product.objects.create(
        supplier=supplier,
        supplier_product_code="SUP-CODE-1",
        sku="SKU-C-1",
        slug="sku-c-1",
        title="P1",
        brand=brand,
        category=category,
        condition=condition,
    )
    Product.objects.create(
        supplier=supplier,
        supplier_product_code="SUP-CODE-2",
        sku="SKU-C-2",
        slug="sku-c-2",
        title="P2",
        brand=brand,
        category=category,
        condition=condition,
    )
    import_file = build_import_file(
        headers=list(CANONICAL_IMPORT_COLUMNS),
        rows=[
            canonical_row(
                product_title="Conflict row",
                brand_name="Brand C",
                category_slug="category-c",
                condition_code="new",
                oe_code="SKU-C-1",
                supplier_product_code="SUP-CODE-2",
            )
        ],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)
    row = result.rows.get(row_number=2)

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED_WITH_ERRORS
    assert result.failed_rows == 1
    assert "different products" in row.error_message


@pytest.mark.django_db
def test_import_avoids_duplicate_brand_and_category_from_case_variants(django_user_model) -> None:
    supplier = make_supplier("SUP-NODUP")
    user = make_staff_user(django_user_model, "import_user_nodup")
    Brand.objects.create(name="Bosch", slug="bosch")
    Category.objects.create(name="Alternator", slug="alternator")
    make_condition("new", "Nuevo", "new")
    import_file = build_import_file(
        headers=list(CANONICAL_IMPORT_COLUMNS),
        rows=[
            canonical_row(
                product_title="Case variant row",
                brand_name="bosch",
                category_slug="alternator",
                condition_code="new",
                oe_code="SKU-NODUP-1",
                supplier_product_code="SUP-NODUP-1",
            )
        ],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED
    assert Brand.objects.filter(name__iexact="bosch").count() == 1
    assert Category.objects.filter(name__iexact="alternator").count() == 1


@pytest.mark.django_db
def test_unknown_extra_columns_are_ignored_with_warning(django_user_model) -> None:
    supplier = make_supplier("SUP-WARN")
    user = make_staff_user(django_user_model, "import_user_warn")
    make_condition("new", "Nuevo", "new")
    Brand.objects.create(name="Brand Warn", slug="brand-warn")
    Category.objects.create(name="Category Warn", slug="category-warn")
    headers = list(CANONICAL_IMPORT_COLUMNS) + ["unexpected_column"]
    row = canonical_row(
        product_title="Warn row",
        brand_name="Brand Warn",
        category_slug="category-warn",
        condition_code="new",
        oe_code="SKU-WARN-1",
        supplier_product_code="SUP-WARN-1",
    )
    import_file = build_import_file(
        headers=headers,
        rows=[[*row, "ignored value"]],
    )
    import_record = SupplierImport.objects.create(
        supplier=supplier,
        uploaded_by=user,
        original_file=import_file,
    )

    result = run_supplier_import(import_record, user)

    assert result.import_status == SupplierImport.ImportStatus.COMPLETED_WITH_ERRORS
    assert "Unknown columns ignored" in result.processing_notes
    assert Product.objects.filter(sku="SKU-WARN-1").exists()


@pytest.mark.django_db
def test_supplier_import_admin_processing_action_authorization(django_user_model) -> None:
    admin_instance = SupplierImportAdmin(SupplierImport, AdminSite())
    internal_user = make_staff_user(django_user_model, "internal_processor")
    internal_user.groups.add(Group.objects.get(name=ROLE_INTERNAL_STAFF))
    restricted_user = make_staff_user(django_user_model, "restricted_processor")
    restricted_user.groups.add(Group.objects.get(name=ROLE_RESTRICTED_SUPPLIER))

    internal_request = RequestFactory().get("/admin/imports/supplierimport/")
    internal_request.user = internal_user
    restricted_request = RequestFactory().get("/admin/imports/supplierimport/")
    restricted_request.user = restricted_user

    internal_actions = admin_instance.get_actions(internal_request)
    restricted_actions = admin_instance.get_actions(restricted_request)

    assert "process_selected_imports" in internal_actions
    assert "process_selected_imports" not in restricted_actions


@pytest.mark.django_db
def test_import_admin_links_to_only_the_products_in_its_batch(django_user_model) -> None:
    supplier = make_supplier("SUP-REVIEW")
    user = make_staff_user(django_user_model, "import_review_user")
    category = Category.objects.create(name="Review", slug="review")
    condition = make_condition("review-new", "Nuevo", "review-new")
    included = Product.objects.create(
        supplier=supplier,
        sku="REVIEW-INCLUDED",
        title="Included",
        category=category,
        condition=condition,
    )
    excluded = Product.objects.create(
        supplier=supplier,
        sku="REVIEW-EXCLUDED",
        title="Excluded",
        category=category,
        condition=condition,
    )
    record = SupplierImport.objects.create(supplier=supplier, uploaded_by=user)
    SupplierImportRow.objects.create(
        supplier_import=record,
        row_number=2,
        processing_status=SupplierImportRow.ProcessingStatus.SUCCESS,
        linked_product=included,
    )

    import_admin = SupplierImportAdmin(SupplierImport, AdminSite())
    link = str(import_admin.review_products_link(record))
    request = RequestFactory().get("/admin/catalog/product/", {"supplier_import": str(record.pk)})
    request.user = user
    params = {"supplier_import": [str(record.pk)]}
    product_filter = ProductAdmin.SupplierImportFilter(
        request, params, Product, ProductAdmin(Product, AdminSite())
    )
    queryset = product_filter.queryset(request, Product.objects.all())

    assert f"supplier_import={record.pk}" in link
    assert list(queryset) == [included]
    assert excluded not in queryset
    assert "original_file" in import_admin.get_readonly_fields(request, record)
